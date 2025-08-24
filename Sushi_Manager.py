import os
import sys
import threading
import time
import re
import copy
import csv
import traceback
from datetime import datetime, timedelta
from collections import defaultdict
from dateutil.parser import parse
import customtkinter as ctk
from tkinter import filedialog, messagebox, simpledialog
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
import codecs
import pandas as pd
from PIL import Image, ImageDraw, ImageFont
from openpyxl.cell.cell import MergedCell
import glob
import xlwings as xw
import tkinter.ttk as ttk
import tempfile
import uuid
import win32com.client
import base64
import pywintypes
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from calendar import monthcalendar

def normalize_supplier_name(name):
    return name.lower().replace(" ", "").replace("(", "").replace(")", "").replace(".", "").replace("_", "")

def get_week_of_month(dt):
    """計算日期在當月是第幾週（從1開始）"""
    # 使用更簡單的邏輯：直接計算日期在當月是第幾週
    # 1-7日為第1週，8-14日為第2週，以此類推
    week_number = (dt.day - 1) // 7 + 1
    return week_number

def find_supplier_file(supplier_name, files):
    norm_supplier = normalize_supplier_name(supplier_name)
    candidates = []
    for f in files:
        # 只要檔名有供應商名稱就算
        norm_f = normalize_supplier_name(os.path.splitext(f)[0])
        if norm_supplier in norm_f:
            candidates.append(f)
    if len(candidates) == 1:
        return candidates[0]
    elif len(candidates) > 1:
        for f in candidates:
            if f.startswith(supplier_name):
                return f
        return candidates[0]
    return None

# ======== 资源路径处理 ========
def resource_path(relative_path):
    """获取资源文件的绝对路径"""
    try:
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# ========== 全局配置 ==========
LOGO_PATH = resource_path("SELOGO22 - 01.png")
PASSWORD = "OPS123"
VERSION = "6.6.0"
DEVELOPER = "OPS - Voon Kee"

# 主题配置
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

# 颜色定义
DARK_BG = "#0f172a"
DARK_PANEL = "#1e293b"
ACCENT_BLUE = "#3b82f6"
BTN_HOVER = "#60a5fa"
ACCENT_GREEN = "#10b981"
ACCENT_RED = "#ef4444"
ACCENT_PURPLE = "#8b5cf6"
ENTRY_BG = "#334155"
TEXT_COLOR = "#e2e8f0"
PANEL_BG = "#1e293b"
TEXTBOX_BG = "#0f172a"

# 字体配置
FONT_TITLE = ("Microsoft JhengHei", 24, "bold")
FONT_BIGBTN = ("Microsoft JhengHei", 16, "bold")
FONT_MID = ("Microsoft JhengHei", 14)
FONT_SUB = ("Microsoft JhengHei", 12)
FONT_ZH = ("Microsoft JhengHei", 12)
FONT_EN = ("Segoe UI", 11, "italic")
FONT_LOG = ("Consolas", 14)

# ========== 多语言支持 ==========
def t(text):
    translations = {
        "mapping_not_available": "分店供应商对应数据不可用\nOutlet-supplier mapping not available",
        "log_not_available": "日志数据不可用\nLog data not available",
        "info": "信息\nInformation",
        "processing": "處理中...\nProcessing...",
        "please_wait": "請稍候...\nPlease wait...",
        "error": "錯誤\nError",
        "login": "系統登錄\nSystem Login",
        "password": "輸入密碼...\nEnter password...",
        "login_btn": "登入\nLogin",
        "exit_confirm": "確定要退出應用程序嗎？\nAre you sure you want to exit the application?",
        "incorrect_pw": "密碼不正確，請重試\nIncorrect password, please try again",
        "main_title": "Sushi Express 自動化工具\nSushi Express Automation Tool",
        "select_function": "請選擇要執行的功能\nPlease select a function",
        "download_title": "Outlook 訂單下載\nOutlook Order Download",
        "download_desc": "下載本週的 Weekly Order 附件\nDownload weekly order attachments",
        "browse": "瀏覽...\nBrowse...",
        "start_download": "開始下載\nStart Download",
        "back_to_menu": "返回主菜單\nBack to Main Menu",
        "checklist_title": "Weekly Order 檢查表\nWeekly Order Checklist",
        "checklist_desc": "請選擇包含供應商訂單的資料夾\nSelect folder with supplier orders",
        "run_check": "執行檢查\nRun Check",
        "automation_title": "訂單自動整合\nOrder Automation",
        "automation_desc": "請選擇三個必要的資料夾\nSelect required folders",
        "source_folder": "來源資料夾 (Weekly Orders)\nSource Folder (Weekly Orders)",
        "supplier_folder": "供應商資料夾 (Supplier)\nSupplier Folder",
        "outlet_folder": "分店資料夾 (Outlet)\nOutlet Folder",
        "start_automation": "開始整合檔案\nStart Automation",
        "processing_orders": "處理訂單\nProcessing Orders",
        "outlet_suppliers": "分店供應商對應\nOutlet-Supplier Mapping",
        "exit_system": "退出系統\nExit System",
        "select_account": "選擇 Outlook 帳號\nSelect Outlook Account",
        "enter_index": "請輸入序號：\nPlease enter index:",
        "download_summary": "下載摘要\nDownload Summary",
        "auto_download": "自動下載\nAuto Downloaded",
        "skipped": "跳過\nSkipped",
        "saved_to": "保存到\nSaved to",
        "check_results": "檢查結果\nCheck Results",
        "success": "成功\nSuccess",
        "warning": "警告\nWarning",
        "folder_warning": "請先選擇所有必要的資料夾\nPlease select all required folders",
        "close": "關閉\nClose",
        "order_processing": "訂單處理進度\nOrder Processing Progress",
        "outlet_supplier_mapping": "分店-供應商對應關係\nOutlet-Supplier Mapping",
        "select_folder": "選擇文件夾\nSelect Folder",
        "view_mapping": "查看分店供應商對應\nView Outlet-Supplier Mapping",
        "view_log": "查看完整日誌\nView Full Log",
        "supplier_files": "已處理的供應商文件\nProcessed Supplier Files",
        "outlet_files": "已處理的分店文件\nProcessed Outlet Files",
        "outlet_orders": "分店訂購情況\nOutlet Orders",
        "supplier_orders": "供應商訂購情況\nSupplier Orders",
        "send_emails": "發送郵件\nSend Emails",
        "operation_supplies": "營運用品\nOperation Supplies",
    }
    return translations.get(text, text)

def get_contrast_color(bg_color):
    # 簡單亮色/暗色對比
    if isinstance(bg_color, str) and bg_color.startswith('#'):
        r = int(bg_color[1:3], 16)
        g = int(bg_color[3:5], 16)
        b = int(bg_color[5:7], 16)
        luminance = (0.299*r + 0.587*g + 0.114*b)
        return '#000000' if luminance > 186 else '#ffffff'
    return '#ffffff'

# ========== 工具函数 ==========
def load_image(path, max_size=(400, 130)):
    """安全加载图像"""
    try:
        if not os.path.exists(path):
            img = Image.new('RGB', max_size, color=DARK_BG[:3])
            draw = ImageDraw.Draw(img)
            font = ImageFont.truetype("arial.ttf", 24)
            draw.text((10,10), "Logo Missing", fill="white", font=font)
        else:
            img = Image.open(path)
        img.thumbnail(max_size, Image.LANCZOS)
        return ctk.CTkImage(img, size=img.size)
    except Exception as e:
        print(f"Error loading image: {e}")
        return None

# ========== 自定义UI组件 ==========
class GlowButton(ctk.CTkButton):
    """发光效果按钮（美化版）"""
    def __init__(self, master, text=None, glow_color=ACCENT_BLUE, **kwargs):
        super().__init__(master, text=text, **kwargs)
        self._glow_color = glow_color
        self._setup_style()
        self._bind_events()

    def _setup_style(self):
        self.configure(
            border_width=0,
            fg_color=self._glow_color,
            hover_color=self._adjust_color(self._glow_color, 40),
            text_color=get_contrast_color(self._glow_color),
            corner_radius=22,  # 更圆角
            font=("Microsoft JhengHei", 20, "bold"),  # 更大更粗
            height=70,  # 更高
            width=340,  # 更宽
            anchor="center"
        )

    def _bind_events(self):
        self.bind("<Enter>", self._on_enter)
        self.bind("<Leave>", self._on_leave)

    def _on_enter(self, event=None):
        self.configure(border_width=4, border_color=self._adjust_color(self._glow_color, 60), fg_color=self._adjust_color(self._glow_color, 20))

    def _on_leave(self, event=None):
        self.configure(border_width=0, fg_color=self._glow_color)
    
    @staticmethod
    def _adjust_color(color, amount):
        if isinstance(color, tuple) and len(color) >= 3:
            r, g, b = color[:3]
            adjusted = tuple(min(255, max(0, x + amount)) for x in (r, g, b))
            if len(color) == 4:
                return adjusted + (color[3],)
            return adjusted
        else:
            color = color.lstrip('#')
            rgb = tuple(int(color[i:i+2], 16) for i in (0, 2, 4))
            adjusted = tuple(min(255, max(0, x + amount)) for x in rgb)
            return f"#{adjusted[0]:02x}{adjusted[1]:02x}{adjusted[2]:02x}"

class ProgressPopup(ctk.CTkToplevel):
    """进度显示弹窗"""
    def __init__(self, parent, title, start_date=None, end_date=None, outlet_count=None):
        super().__init__(parent)
        self.title(title)
        self.geometry("900x700")
        self.transient(parent)
        self.grab_set()
        self.parent = parent
        self.configure(fg_color=DARK_BG)
        self.log_text = ctk.CTkTextbox(
            self,
            wrap="word",
            font=FONT_LOG,
            fg_color=TEXTBOX_BG,
            text_color=TEXT_COLOR,
            corner_radius=10
        )
        self.log_text.pack(fill="both", expand=True, padx=20, pady=20)
        self.log_text.configure(state="normal")
        if start_date and end_date:
            self.log_text.insert("end", f"📅 抓取日期範圍: {start_date} ~ {end_date}\n")
        self.log_text.configure(state="disabled")
        self.outlet_count_label = ctk.CTkLabel(
            self,
            text=f"🏪 已下載分店數量: {outlet_count if outlet_count is not None else 0} 間",
            font=("Microsoft JhengHei", 18, "bold"),
            text_color="#10b981"
        )
        self.outlet_count_label.pack(pady=(0, 10))
        close_btn = GlowButton(
            self,
            text=t("close"),
            command=self.destroy_popup,
            width=120,
            height=40
        )
        close_btn.pack(pady=10)
    def update_outlet_count(self, count):
        if self.outlet_count_label:
            self.outlet_count_label.configure(text=f"🏪 已下載分店數量: {count} 間")
    def destroy_popup(self):
        self.destroy()
        self.parent.progress_popup = None
    def log(self, message):
        self.log_text.configure(state="normal")
        self.log_text.insert("end", message)
        self.log_text.see("end")
        self.log_text.configure(state="disabled")
        # 強制更新UI
        self.update()

class MappingPopup(ctk.CTkToplevel):
    """分店-供应商映射显示"""
    def __init__(self, parent, title):
        super().__init__(parent)
        self.title(title)
        self.geometry("900x700")
        self.transient(parent)
        self.grab_set()
        self.parent = parent
        self.configure(fg_color=DARK_BG)
        
        self.mapping_text = ctk.CTkTextbox(
            self,
            wrap="word",
            font=FONT_LOG,
            fg_color=TEXTBOX_BG,
            text_color=TEXT_COLOR,
            corner_radius=10
        )
        self.mapping_text.pack(fill="both", expand=True, padx=20, pady=20)
        self.mapping_text.configure(state="disabled")
        
        close_btn = GlowButton(
            self,
            text=t("close"),
            command=self.destroy_popup,
            width=120,
            height=40
        )
        close_btn.pack(pady=10)
    
    def destroy_popup(self):
        self.destroy()
        self.parent.mapping_popup = None
    
    def update_mapping(self, mapping):
        self.mapping_text.configure(state="normal")
        self.mapping_text.delete("1.0", "end")
        self.mapping_text.insert("1.0", mapping)
        self.mapping_text.configure(state="disabled")

class ScrollableMessageBox(ctk.CTkToplevel):
    """可滚动消息框"""
    def __init__(self, parent, title, message):
        super().__init__(parent)
        self.title(title)
        self.geometry("900x700")
        self.transient(parent)
        self.grab_set()
        self.configure(fg_color=DARK_BG)
        
        self.text_box = ctk.CTkTextbox(
            self,
            wrap="word",
            font=FONT_LOG,
            fg_color=TEXTBOX_BG,
            text_color=TEXT_COLOR,
            corner_radius=10
        )
        self.text_box.pack(fill="both", expand=True, padx=20, pady=20)
        self.text_box.insert("1.0", message)
        self.text_box.configure(state="disabled")
        
        close_btn = GlowButton(
            self,
            text=t("close"),
            command=self.destroy,
            width=120,
            height=40
        )
        close_btn.pack(pady=10)

class EmailConfirmationDialog(ctk.CTkToplevel):
    """邮件发送确认对话框（只顯示純文字，保留空行）"""
    def __init__(self, parent, mail_item, supplier_name, outlet_name, attachment_path, on_confirm):
        super().__init__(parent)
        self.title(f"确认邮件 - {supplier_name}")
        self.geometry("800x700")
        self.transient(parent)
        self.grab_set()
        self.mail_item = mail_item
        self.on_confirm = on_confirm
        self.attachment_path = attachment_path
        self.configure(fg_color=DARK_BG)
        info_frame = ctk.CTkFrame(self, fg_color=DARK_PANEL, corner_radius=12)
        info_frame.pack(fill="x", padx=20, pady=10)
        ctk.CTkLabel(info_frame, text=f"收件人(To)/To: {mail_item.To}", font=FONT_MID).pack(anchor="w", padx=10, pady=5)
        ctk.CTkLabel(info_frame, text=f"抄送(CC)/CC: {mail_item.CC}", font=FONT_MID).pack(anchor="w", padx=10, pady=5)
        ctk.CTkLabel(info_frame, text=f"主题/Subject: {mail_item.Subject}", font=FONT_MID).pack(anchor="w", padx=10, pady=5)
        ctk.CTkLabel(info_frame, text=f"供应商/Supplier: {supplier_name}", font=FONT_MID).pack(anchor="w", padx=10, pady=5)
        ctk.CTkLabel(info_frame, text=f"分店/Outlet: {outlet_name}", font=FONT_MID).pack(anchor="w", padx=10, pady=5)
        if attachment_path:
            file_name = os.path.basename(attachment_path)
            ctk.CTkLabel(info_frame, text=f"附件/Attachment: {file_name}", font=FONT_MID).pack(anchor="w", padx=10, pady=5)
        body_frame = ctk.CTkFrame(self, fg_color=DARK_PANEL, corner_radius=12)
        body_frame.pack(fill="both", expand=True, padx=20, pady=10)
        ctk.CTkLabel(body_frame, text="邮件正文/Email Body:", font=FONT_BIGBTN).pack(anchor="w", padx=10, pady=5)
        # 只顯示純文字，保留空行，避免名字被拆行
        import html
        raw_html = mail_item.HTMLBody or ""
        import re
        # 1. 只把結束分行標籤換成換行
        text = re.sub(r'(<br\s*/?>|</div>|</p>|</li>|</tr>|</td>)', '\n', raw_html, flags=re.IGNORECASE)
        # 2. 把開始標籤直接移除（不換行）
        text = re.sub(r'(<div>|<p>|<li>|<tr>|<td>)', '', text, flags=re.IGNORECASE)
        # 3. 移除其他 HTML 標籤
        text = re.sub(r'<[^>]+>', '', text)
        # 4. decode HTML entity
        text = html.unescape(text)
        # 5. 移除開頭/結尾多餘空白
        text = text.strip()
        self.body_text = ctk.CTkTextbox(body_frame, wrap="word", height=300, font=FONT_MID)
        self.body_text.pack(fill="both", expand=True, padx=10, pady=5)
        self.body_text.insert("1.0", text)
        self.body_text.configure(state="disabled")
        btn_frame = ctk.CTkFrame(self, fg_color="transparent")
        btn_frame.pack(fill="x", padx=20, pady=10)
        ctk.CTkButton(
            btn_frame, 
            text="发送邮件/Send Email", 
            command=self._send_email,
            fg_color=ACCENT_GREEN,
            hover_color=BTN_HOVER
        ).pack(side="right", padx=10)
        ctk.CTkButton(
            btn_frame, 
            text="取消/Cancel", 
            command=self.destroy
        ).pack(side="right", padx=10)
        ctk.CTkButton(
            btn_frame, 
            text="编辑正文/Edit Body", 
            command=self._edit_body,
            fg_color=ACCENT_BLUE,
            hover_color=BTN_HOVER
        ).pack(side="left", padx=10)
    def _edit_body(self):
        self.body_text.configure(state="normal")
    def _send_email(self):
        try:
            # 1. 读取文本框内容
            body = self.body_text.get("1.0", "end-1c")

            # 2. 计算"月+周"英文字符串
            today = datetime.now()
            # 获取完整英文月份名称，如 "July"
            month_name = today.strftime("%B")
            # 使用正確的週數計算方法
            week_of_month = get_week_of_month(today)
            # 构造 "July Week 3"
            month_week_str = f"{month_name} Week {week_of_month}"

            # 3. 在模板中替换占位符
            body = body.replace("{week_no}", month_week_str)

            # 4. 转成 HTML 并发送
            html_body = body.replace("\n", "<br>")
            self.mail_item.HTMLBody = html_body
            self.mail_item.Send()
            self.on_confirm(True, "邮件发送成功！")
        except Exception as e:
            self.on_confirm(False, f"发送失败：{e}")
        finally:
            self.destroy()

class NavigationButton(ctk.CTkButton):
    """自定义导航按钮，支持选中状态"""
    def __init__(self, master, text, command, **kwargs):
        super().__init__(master, text=text, command=command, **kwargs)
        self.default_color = DARK_PANEL
        self.selected_color = ACCENT_BLUE
        self.hover_color = BTN_HOVER
        self.is_selected = False
        self.default_border_color = ACCENT_BLUE
        self.selected_border_color = ACCENT_GREEN
        self.configure(
            fg_color=self.default_color,
            hover_color=self.hover_color,
            text_color=get_contrast_color(self.default_color),
            border_width=2,
            border_color=self.default_border_color,
            corner_radius=10,
            font=FONT_BIGBTN,
            height=50,
            anchor="center"
        )
        self._bind_events()
    def _bind_events(self):
        self.bind("<Enter>", self._on_enter)
        self.bind("<Leave>", self._on_leave)
    def _on_enter(self, event=None):
        if not self.is_selected:
            self.configure(fg_color=self._adjust_color(self.default_color, 20))
    def _on_leave(self, event=None):
        if not self.is_selected:
            self.configure(fg_color=self.default_color)
    def select(self):
        self.is_selected = True
        self.configure(
            fg_color=self.selected_color,
            border_width=2,
            border_color=self.selected_border_color,
            text_color=get_contrast_color(self.selected_color)
        )
    def deselect(self):
        self.is_selected = False
        self.configure(
            fg_color=self.default_color,
            border_width=2,
            border_color=self.default_border_color,
            text_color=get_contrast_color(self.default_color)
        )
    @staticmethod
    def _adjust_color(color, amount):
        if isinstance(color, str) and color.startswith("#"):
            color = color.lstrip('#')
            rgb = tuple(int(color[i:i+2], 16) for i in (0, 2, 4))
            adjusted = tuple(min(255, max(0, x + amount)) for x in rgb)
            return f"#{adjusted[0]:02x}{adjusted[1]:02x}{adjusted[2]:02x}"
        return color

# ========== 送货日期验证工具 ==========
class DeliveryDateValidator:
    """送货日期验证工具"""
    
    DAYS_MAPPING = {
        'mon': 0, 'monday': 0, '星期一': 0,
        'tue': 1, 'tuesday': 1, '星期二': 1,
        'wed': 2, 'wednesday': 2, '星期三': 2,
        'thu': 3, 'thursday': 3, '星期四': 3,
        'fri': 4, 'friday': 4, '星期五': 4,
        'sat': 5, 'saturday': 5, '星期六': 5,
        'sun': 6, 'sunday': 6, '星期日': 6
    }

    def __init__(self, config_file=None):
        self.schedule = defaultdict(dict)
        if config_file:
            self.load_config(config_file)
    
    def load_config(self, config_file):
        """加载送货日期配置"""
        try:
            with open(config_file, mode='r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    supplier = row['supplier'].strip().upper()
                    outlet = row['outlet_code'].strip().upper()
                    days = self.parse_delivery_days(row['delivery_days'])
                    
                    if outlet == "ALL":
                        self.schedule[supplier]['*'] = days
                    else:
                        self.schedule[supplier][outlet] = days
        except Exception as e:
            raise Exception(f"加载送货配置失败: {str(e)}")

    def parse_delivery_days(self, days_str):
        """解析送货日期字符串"""
        days = set()
        for day in days_str.split(','):
            day = day.strip().lower()
            if day in self.DAYS_MAPPING:
                days.add(self.DAYS_MAPPING[day])
        return days

    def get_delivery_days(self, supplier, outlet_code):
        supplier = supplier.strip().upper()
        outlet_code = outlet_code.strip().upper()
        outlet_specific = self.schedule.get(supplier, {}).get(outlet_code)
        if outlet_specific is not None:
            return outlet_specific
        return self.schedule.get(supplier, {}).get('*', set())

    def validate_order(self, supplier, outlet_code, order_date, log_callback=None):
        supplier = supplier.strip().upper()
        outlet_code = outlet_code.strip().upper()
        delivery_days = self.get_delivery_days(supplier, outlet_code)
        
        if not delivery_days:
            return True
        
        try:
            if isinstance(order_date, (int, float)):
                order_date = datetime(1899, 12, 30) + timedelta(days=order_date)
            else:
                order_date = parse(str(order_date), fuzzy=True)
            
            order_weekday = order_date.weekday()
            
            if order_weekday not in delivery_days:
                day_name = ['Monday', 'Tuesday', 'Wednesday', 'Thursday',
                           'Friday', 'Saturday', 'Sunday'][order_weekday]
                if log_callback:
                    log_callback(
                        f"❌ 送货日期错误: {outlet_code} 向 {supplier} 下单\n"
                        f"  订单日期: {order_date.strftime('%Y-%m-%d')} ({day_name})\n"
                        f"  允许送货日: {self.format_days(delivery_days)}"
                    )
                return False
            return True
        except Exception as e:
            if log_callback:
                log_callback(f"⚠️ 日期解析失败: {supplier}-{outlet_code} ({order_date}): {str(e)}")
            return False

    def format_days(self, day_numbers):
        """将数字转换为星期名称"""
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 
               'Friday', 'Saturday', 'Sunday']
        return ", ".join(days[d] for d in sorted(day_numbers))

# ========== 统一配置管理器 ==========
class UnifiedConfigManager:
    """管理统一的Excel配置文件"""
    
    def __init__(self, config_path=None):
        self.config_path = config_path
        self.outlets = []
        self.suppliers = []
        self.delivery_schedule = []
        self.email_templates = {}
        self.supplier_requirements = {}
        
        if config_path:
            self.load_config(config_path)
    
    def load_config(self, config_path):
        """从Excel文件加载配置"""
        try:
            wb = load_workbook(config_path, data_only=True)
            
            # 读取分店信息
            if "Outlets" in wb.sheetnames:
                ws = wb["Outlets"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and row[0]:
                        self.outlets.append({
                            "code": row[0].strip().upper() if isinstance(row[0], str) else str(row[0]).strip().upper(),
                            "name": row[1].strip() if len(row) > 1 and isinstance(row[1], str) else str(row[1]).strip() if len(row) > 1 and row[1] is not None else "",
                            "email": row[2].strip().lower() if len(row) > 2 and isinstance(row[2], str) else str(row[2]).strip().lower() if len(row) > 2 and row[2] is not None else "",
                            "address": row[3].strip() if len(row) > 3 and isinstance(row[3], str) else str(row[3]).strip() if len(row) > 3 and row[3] is not None else "",
                            "delivery_day": row[4].strip() if len(row) > 4 and isinstance(row[4], str) else str(row[4]).strip() if len(row) > 4 and row[4] is not None else "",
                            "brand": row[5].strip() if len(row) > 5 and isinstance(row[5], str) else str(row[5]).strip() if len(row) > 5 and row[5] is not None else "Dine-In"
                        })
            
            # 读取供应商信息
            if "Suppliers" in wb.sheetnames:
                ws = wb["Suppliers"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and row[0]:
                        self.suppliers.append({
                            "name": row[0].strip() if isinstance(row[0], str) else str(row[0]).strip(),
                            "email": row[1].strip().lower() if len(row) > 1 and isinstance(row[1], str) else str(row[1]).strip().lower() if len(row) > 1 and row[1] is not None else "",
                            "contact": row[2].strip() if len(row) > 2 and isinstance(row[2], str) else str(row[2]).strip() if len(row) > 2 and row[2] is not None else "",
                            "phone": row[3].strip() if len(row) > 3 and isinstance(row[3], str) else str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
                        })
            
            # 读取配送日程
            if "Delivery Schedule" in wb.sheetnames:
                ws = wb["Delivery Schedule"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and row[0]:
                        self.delivery_schedule.append({
                            "supplier": row[0].strip() if isinstance(row[0], str) else str(row[0]).strip(),
                            "outlet_code": row[1].strip().upper() if len(row) > 1 and isinstance(row[1], str) else str(row[1]).strip().upper() if len(row) > 1 and row[1] is not None else "ALL",
                            "delivery_days": row[2].strip() if len(row) > 2 and isinstance(row[2], str) else str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
                        })
            
            # 读取邮件模板
            if "Email Templates" in wb.sheetnames:
                ws = wb["Email Templates"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and len(row) > 0 and row[0]:
                        self.email_templates[row[0].strip() if isinstance(row[0], str) else str(row[0]).strip()] = {
                            "subject": row[1].strip() if len(row) > 1 and isinstance(row[1], str) else str(row[1]).strip() if len(row) > 1 and row[1] is not None else "",
                            "body": row[2].strip() if len(row) > 2 and isinstance(row[2], str) else str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
                        }
            
            # 读取供应商要求
            if "Supplier Requirements" in wb.sheetnames:
                ws = wb["Supplier Requirements"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and row[0]:
                        supplier_name = row[0].strip() if isinstance(row[0], str) else str(row[0]).strip()
                        # 修正：正確分割、去空格、轉大寫
                        outlet_codes = []
                        if len(row) > 1 and row[1] is not None:
                            for code in str(row[1]).replace("\n", ",").replace("，", ",").split(","):
                                code = code.strip().upper()
                                if code:
                                    outlet_codes.append(code)
                        self.supplier_requirements[supplier_name] = outlet_codes
            
            return True, f"成功加载配置文件: {len(self.outlets)} 分店, {len(self.suppliers)} 供应商"
        except Exception as e:
            return False, f"加载配置文件失败: {str(e)}"
    
    def get_outlet(self, code):
        # 根據分店代碼獲取分店全名
        if not code:
            return None
        if isinstance(code, str):
            code = code.strip().upper()
        else:
            code = str(code).upper()
        for outlet in self.outlets:
            if outlet['short_name'].upper() == code:
                return outlet['full_name']
        return None
    
    def get_supplier(self, name):
        # 根據供應商名稱獲取標準名稱
        if not name:
            return None
        if isinstance(name, str):
            name = name.strip().upper()
        else:
            name = str(name).upper()
        for supplier in self.suppliers:
            if supplier['name'].upper() == name:
                return supplier['name']
        return None
    
    def get_delivery_schedule(self, supplier, outlet_code):
        """获取特定供应商-分店的配送日程"""
        for schedule in self.delivery_schedule:
            if schedule["supplier"] == supplier and schedule["outlet_code"] == outlet_code:
                return schedule["delivery_days"]
        
        for schedule in self.delivery_schedule:
            if schedule["supplier"] == supplier and schedule["outlet_code"] == "ALL":
                return schedule["delivery_days"]
        
        return None
    
    def get_required_outlets(self, supplier_name):
        # 用 normalize 方式找 key，确保匹配
        key = str(supplier_name).strip().upper()
        for k in self.supplier_requirements:
            if str(k).strip().upper() == key:
                return self.supplier_requirements[k]
        return []

# ========== 邮件发送管理器 ==========
class EmailSender:
    """处理邮件发送功能"""

    def __init__(self, config_manager=None):
        self.config_manager = config_manager
        # 加载 GIF 图片资源
        self.email_gif = self._load_email_gif()

    def _load_email_gif(self):
        """加载邮件签名 GIF 图片"""
        try:
            gif_path = resource_path("email_gif.gif")
            if os.path.exists(gif_path):
                with open(gif_path, "rb") as f:
                    return f.read()
            return None
        except Exception as e:
            print(f"Error loading email GIF: {str(e)}")
            return None

    def _get_standard_subject(self, supplier_name):
        """生成标准邮件主题"""
        now = datetime.now()
        month_name = now.strftime("%B")
        week_in_month = get_week_of_month(now)
        return f"Sushi Express Weekly Order - {supplier_name} - {month_name} - Week {week_in_month}"

    def get_to_cc_emails(self, supplier_name, config_path):
        """根據 config excel 取得 TO/CC 郵件，並自動加 opsadmin 及 purchasing.admin，且CC不重複"""
        wb = load_workbook(config_path, data_only=True)
        to_emails = []
        cc_emails = []
        if "Suppliers" in wb.sheetnames:
            ws = wb["Suppliers"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                if str(row[0]).strip() == supplier_name:
                    typ = str(row[1]).strip().upper() if len(row) > 1 and row[1] else "TO"
                    email = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                    if not email:
                        continue
                    if typ == "TO":
                        to_emails.append(email)
                    elif typ == "CC":
                        cc_emails.append(email)
        # 必須CC的名單
        must_cc = [
            "opsadmin@sushiexpress.com.sg",
            "purchasing.admin@sushiexpress.com.sg"
        ]
        for cc in must_cc:
            if cc not in cc_emails:
                cc_emails.append(cc)
        # 去重（不分大小寫）
        seen = set()
        cc_unique = []
        for e in cc_emails:
            elower = e.lower()
            if elower not in seen:
                cc_unique.append(e)
                seen.add(elower)
        return to_emails, cc_unique

    def send_email(self, to_emails, cc_emails, supplier_name, body, attachment_path=None, account_idx=None, use_content_id=True, subject=None):
        import time
        import pywintypes
        max_retries = 3
        for attempt in range(max_retries):
            try:
                import win32com.client
                import os
                import base64
                outlook = win32com.client.Dispatch("Outlook.Application")
                # --- 新增 CreateItem 自動重試 ---
                for create_attempt in range(3):
                    try:
                        mail = outlook.CreateItem(0)
                        break
                    except AttributeError as e:
                        if "CreateItem" in str(e):
                            print(f"[RETRY] Outlook COM 還沒準備好，第 {create_attempt+1} 次重試...")
                            time.sleep(2)
                            outlook = win32com.client.Dispatch("Outlook.Application")
                            continue
                        else:
                            raise
                else:
                    from tkinter import messagebox
                    messagebox.showerror("邮件发送失败", "Outlook 啟動異常，請確認 Outlook 已開啟且無彈窗。")
                    return None, "Outlook 啟動異常，請確認 Outlook 已開啟且無彈窗。"
                # --- 其餘原本的 send_email 流程 ---
                if account_idx is not None:
                    mapi = outlook.GetNamespace("MAPI")
                    accounts = [mapi.Folders.Item(i + 1) for i in range(mapi.Folders.Count)]
                    if 0 <= account_idx < len(accounts):
                        account = accounts[account_idx]
                        if hasattr(mail, 'SendUsingAccount'):
                            for acc in outlook.Session.Accounts:
                                if acc.DisplayName == account.Name:
                                    mail.SendUsingAccount = acc
                                    break
                # 處理收件人
                if isinstance(to_emails, list):
                    to_emails_fixed = ';'.join(to_emails) if to_emails else ''
                else:
                    to_emails_fixed = to_emails.replace('，', ';').replace(',', ';') if to_emails else ''
                if isinstance(cc_emails, list):
                    cc_emails_fixed = ';'.join(cc_emails) if cc_emails else ''
                else:
                    cc_emails_fixed = cc_emails.replace('，', ';').replace(',', ';') if cc_emails else ''
                mail.To = to_emails_fixed
                mail.CC = cc_emails_fixed
                if subject:
                    mail.Subject = subject
                else:
                    mail.Subject = self._get_standard_subject(supplier_name)
                # 添加 GIF 簽名
                possible_gif_paths = [
                    "email_gif.gif",
                    os.path.join(os.path.dirname(__file__), "email_gif.gif"),
                    os.path.join(os.getcwd(), "email_gif.gif"),
                    os.path.abspath("email_gif.gif")
                ]
                signature_gif_path = None
                for path in possible_gif_paths:
                    if os.path.exists(path):
                        signature_gif_path = path
                        break
                print(f"[DEBUG] 檢查 GIF 文件路徑: {possible_gif_paths}")
                print(f"[DEBUG] 找到的 GIF 文件: {signature_gif_path}")
                print(f"[DEBUG] 當前工作目錄: {os.getcwd()}")
                print(f"[DEBUG] 腳本目錄: {os.path.dirname(__file__)}")
                html_body = body
                if not any(tag in body.lower() for tag in ['<br', '<p', '<div', '<table', '<ul', '<ol', '<li', '<b', '<strong', '<em', '<span']):
                    html_body = body.replace('\n', '<br>')
                if signature_gif_path and os.path.exists(signature_gif_path):
                    if use_content_id:
                        # Content-ID 方式（新版 Outlook 兼容）
                        cid = "sigimg001"
                        signature_html = f"""
                        <br><br>
                        <img src=\"cid:{cid}\" alt=\"Signature\" style=\"max-width: 400px;\">
                        """
                        html_body += signature_html
                        mail.HTMLBody = html_body
                        att = mail.Attachments.Add(os.path.abspath(signature_gif_path))
                        att.PropertyAccessor.SetProperty("http://schemas.microsoft.com/mapi/proptag/0x3712001F", cid)
                        print(f"[DEBUG] 已用 Content-ID 插入 GIF 簽名")
                    else:
                        # base64 方式（舊版 Outlook 兼容）
                        with open(signature_gif_path, 'rb') as f:
                            gif_data = base64.b64encode(f.read()).decode()
                        signature_html = f"""
                        <br><br>
                        <img src=\"data:image/gif;base64,{gif_data}\" alt=\"Signature\" style=\"max-width: 400px;\">
                        """
                        html_body += signature_html
                        mail.HTMLBody = html_body
                        print(f"[DEBUG] 已用 base64 插入 GIF 簽名")
                else:
                    mail.HTMLBody = html_body
                # 添加订单附件
                if attachment_path and os.path.exists(attachment_path):
                    mail.Attachments.Add(attachment_path)
                return mail
            except pywintypes.com_error as e:
                if hasattr(e, 'args') and len(e.args) > 0 and e.args[0] == -2147418111:
                    print(f"[RETRY] Outlook 忙碌中，第 {attempt+1} 次重試...")
                    time.sleep(2)
                    continue
                else:
                    import traceback
                    print(traceback.format_exc())
                    from tkinter import messagebox
                    messagebox.showerror("邮件发送失败", f"{str(e)}\n\n{traceback.format_exc()}")
                    return None, f"创建邮件失败: {str(e)}"
            except Exception as e:
                import traceback
                print(traceback.format_exc())
                from tkinter import messagebox
                messagebox.showerror("邮件发送失败", f"{str(e)}\n\n{traceback.format_exc()}")
                return None, f"创建邮件失败: {str(e)}"
        # 如果重試後還是失敗
        from tkinter import messagebox
        messagebox.showerror("邮件发送失败", "Outlook 忙碌，重试多次仍失败。请稍后再试。")
        return None, "Outlook 忙碌，重试多次仍失败。请稍后再试。"

# ========== 订单自动化核心 ==========
class OrderAutomation:
    """订单自动化工具"""
    
    def __init__(self, outlet_config=None):
        # outlet_config: list of dicts with keys 'short_name', 'full_name'
        self.outlet_name_map = {}
        if outlet_config:
            for o in outlet_config:
                full = o.get('full_name', '').strip().lower()
                short = o.get('short_name', '').strip().upper()
                if full:
                    self.outlet_name_map[full] = short
                if short:
                    self.outlet_name_map[short.lower()] = short
    def get_short_code(self, f5val):
        val = (str(f5val) or '').strip().lower()
        if val in self.outlet_name_map:
            return self.outlet_name_map[val]
        # 雙向模糊比對
        for full, short in self.outlet_name_map.items():
            if full in val or val in full:
                return short
            # 單字比對
            for word in val.split():
                if word and word in full:
                    return short
        return 'UNKNOWN'

    @staticmethod
    def is_valid_date(cell_value, next_week_start, next_week_end):
        """检查是否为有效日期"""
        try:
            if isinstance(cell_value, (int, float)):
                base_date = datetime(1899, 12, 30)
                parsed = base_date + timedelta(days=cell_value)
            else:
                parsed = parse(str(cell_value), fuzzy=True, dayfirst=False)
            # 對於 Amendment 檔案，我們接受任何日期，不限制在特定週期內
            return True  # 暫時接受所有日期
        except:
            return False

    @classmethod
    def find_delivery_date_row(cls, ws, next_week_start, next_week_end, max_rows=200, file_path=None):
        """查找送货日期行"""
        valid_col_range = range(5, 12)
        invalid_labels = ["total", "total:", "sub-total", "sub-total:", "no. of cartons", "no. of cartons:"]
        found_blocks = []
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_rows)):
            cols = []
            for j, cell in enumerate(row):
                if j not in valid_col_range:
                    continue
                val = cell.value
                if cls.is_valid_date(val, next_week_start, next_week_end):
                    cols.append(j)
            if cols:
                found_blocks.append((i + 1, cols))
        for header_row, cols in found_blocks:
            for row_idx in range(header_row + 1, header_row + 200):
                label_val = ws.cell(row=row_idx, column=5).value
                label = str(label_val).lower().strip() if label_val else ""
                if any(invalid in label for invalid in invalid_labels):
                    continue
                for col_idx in cols:
                    qty_val = ws.cell(row=row_idx, column=col_idx+1).value
                    if isinstance(qty_val, (int, float)) and qty_val > 0:
                        return header_row, cols
        return None, []

    @classmethod
    def run_automation(cls, source_folder, supplier_folder, 
                      outlet_config=None, delivery_config=None,
                      log_callback=None, mapping_callback=None):
        """运行订单自动化（仅生成 supplier 文件）"""
        now = datetime.now()
        today = now.date()
        this_monday = today - timedelta(days=today.weekday())
        next_monday = this_monday + timedelta(days=7)
        next_sunday = next_monday + timedelta(days=6)
        start_of_period = datetime.combine(next_monday, datetime.min.time())
        end_of_period = datetime.combine(next_sunday, datetime.max.time())
        save_path = os.path.join(source_folder, f"next_week_order_log_{now.strftime('%Y%m%d_%H%M%S')}.txt")
        log_lines = [f"🎯 Next Week Order Integration Log\n", f"Scan Start Time: {now}\n", f"Target Week: {start_of_period.date()} to {end_of_period.date()}\n"]
        
        date_validator = DeliveryDateValidator(delivery_config) if delivery_config else None
        
        outlet_mapping = {}
        if outlet_config:
            try:
                outlet_mapping = {o['short_name']: o for o in outlet_config}
                if log_callback:
                    log_callback(f"✅ 已加载分店配置: {len(outlet_mapping)} 个分店")
            except Exception as e:
                if log_callback:
                    log_callback(f"⚠️ 分店配置加载失败: {str(e)}")
        
        # 新增：建立一個 OrderAutomation 實例用於 get_short_code
        oa_instance = cls(outlet_config) if outlet_config else cls()
        
        supplier_to_outlets = defaultdict(list)
        
        def log(message, include_timestamp=True):
            timestamp = datetime.now().strftime("%H:%M:%S") if include_timestamp else ""
            log_line = f"[{timestamp}] {message}" if include_timestamp else message
            log_lines.append(log_line + "\n")
            if log_callback:
                log_callback(log_line + "\n")

        files = [f for f in os.listdir(source_folder)
                 if f.endswith((".xlsx", ".xls")) and not f.startswith("~$")]
        total_files = len(files)

        if total_files == 0:
            log("未找到Excel文件\nNo Excel files found in source folder. Exiting.")
            return False, "在源文件夹中未找到Excel文件"

        log(f"找到 {total_files} 个Excel文件\nFound {total_files} Excel files")
        log(f"目标周期: {start_of_period.strftime('%Y-%m-%d')} 到 {end_of_period.strftime('%Y-%m-%d')}\nTarget period: {start_of_period.strftime('%Y-%m-%d')} to {end_of_period.strftime('%Y-%m-%d')}")
        log(f"🔍 專注於下週訂單，忽略其他週期\n🔍 Focus on next week orders only, ignore other periods")

        for idx, file in enumerate(files):
            full_path = os.path.join(source_folder, file)
            try:
                # log(f"\n处理文件 {idx+1}/{total_files}: {file}\nProcessing file {idx+1}/{total_files}: {file}")
                wb = load_workbook(full_path, data_only=True)
                # log(f"工作表: {', '.join(wb.sheetnames)}\nWorksheets: {', '.join(wb.sheetnames)}")
                for sheetname in wb.sheetnames:
                    ws = wb[sheetname]
                    if ws.sheet_state != "visible":
                        continue  # 跳過隱藏工作表不 log，不顯示任何訊息
                    outlet_short = file.split('_')[0].strip() if '_' in file else file.split('.')[0].strip()
                    # log(f"  工作表: {sheetname}, 文件名分店简称: '{outlet_short}'\n  Sheet: {sheetname}, Short Name (from filename): '{outlet_short}'")
                    has_order = False
                    week_days = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
                    for row in range(1, ws.max_row):
                        row_vals = [str(ws.cell(row=row, column=col).value).strip() if ws.cell(row=row, column=col).value else '' for col in range(6, 13)]
                        weekday_count = sum(day in row_vals for day in week_days)
                        if 'Mon' in row_vals and weekday_count >= 3:
                            date_row = row + 1
                            date_cols = []
                            for idx2, col in enumerate(range(6, 13)):
                                val = ws.cell(row=date_row, column=col).value
                                parsed = None
                                try:
                                    if val:
                                        sval = str(val)
                                        if re.match(r"^\d{1,2}-[A-Za-z]{3,}$", sval):
                                            year = start_of_period.year
                                            sval = f"{sval}-{year}"
                                        if isinstance(val, (int, float)):
                                            base_date = datetime(1899, 12, 30)
                                            parsed = base_date + timedelta(days=val)
                                        else:
                                            parsed = parse(sval, fuzzy=True, dayfirst=False)
                                except Exception:
                                    pass
                                if parsed and start_of_period.date() <= parsed.date() <= end_of_period.date():
                                    date_cols.append(col)
                            if date_cols:
                                for r in range(date_row+1, min(date_row+30, ws.max_row+1)):
                                    found_order_this_row = False
                                    for c in date_cols:
                                        val = ws.cell(row=r, column=c).value
                                        if isinstance(val, (datetime,)):
                                            if val.date() > end_of_period.date():
                                                found_order_this_row = True
                                                break
                                        elif isinstance(val, (int, float)) and val > 0:
                                            has_order = True
                                            found_order_this_row = True
                                            break
                                    if found_order_this_row:
                                        break
                            if has_order:
                                break
                    if has_order:
                        # 只記錄有訂單的廠商與分店，不記錄細節
                        supplier_to_outlets[sheetname].append((outlet_short, full_path, sheetname))
            except Exception as e:
                error_msg = f"❌ 处理文件出错: {file}\n❌ Error processing {file}: {str(e)}\n{traceback.format_exc()}"
                log(error_msg)
        
        log("\n📊 下週訂單整合結果 / Next Week Order Integration Results:")
        for supplier, outlets in supplier_to_outlets.items():
            outlet_list = [str(o[0]) if o[0] else "UNKNOWN" for o in outlets]
            log(f"  📦 {supplier}: {', '.join(outlet_list)}")
        log(f"\n🎯 只處理下週訂單，共 {sum(len(outlets) for outlets in supplier_to_outlets.values())} 個門市有訂單")
        log("\nCreating supplier files...")
        supplier_files = []
        import xlwings as xw
        for sheetname, outlet_file_pairs in supplier_to_outlets.items():
            supplier_path = os.path.join(supplier_folder, f"{sheetname}_Week_{now.isocalendar()[1]}.xlsx")
            # 先建立空檔案
            if not os.path.exists(supplier_path):
                from openpyxl import Workbook
                wb = Workbook()
                wb.save(supplier_path)
            app = xw.App(visible=False)
            try:
                wb_dest = app.books.open(supplier_path)
                for outlet, src_file, original_sheet in outlet_file_pairs:
                    try:
                        wb_src = app.books.open(src_file)
                        sht = wb_src.sheets[original_sheet]
                        dest_sheet_name = outlet if outlet else original_sheet
                        # 刪除同名分頁
                        for s in wb_dest.sheets:
                            if s.name == dest_sheet_name:
                                s.delete()
                        sht.api.Copy(Before=wb_dest.sheets[0].api)
                        wb_dest.sheets[0].name = dest_sheet_name
                        wb_src.close()
                        log(f"  ✅ {dest_sheet_name} 已複製進 {os.path.basename(supplier_path)}（格式/公式完整保留）")
                    except Exception as e:
                        log(f"    ❌ Failed to copy {outlet} in {sheetname}: {str(e)}\n{traceback.format_exc()}")
                wb_dest.save()
                wb_dest.close()
                supplier_files.append(os.path.basename(supplier_path))
            finally:
                app.quit()
        result_text = "\n🎯 下週訂單整合完成 / Next Week Order Integration Complete:\n"
        result_text += f"📦 已處理供應商文件: {len(supplier_files)}\n"
        result_text += f"📅 目標週期: {start_of_period.strftime('%Y-%m-%d')} 到 {end_of_period.strftime('%Y-%m-%d')}\n\n"
        result_text += "📋 供應商文件列表:\n" + "\n".join([f"  ✅ {file}" for file in supplier_files])
        log(result_text)
        try:
            with open(save_path, "w", encoding="utf-8") as logfile:
                logfile.writelines(log_lines)
            log(f"\nLog saved at: {save_path}")
            return True, f"订单整合完成！\n\n日志文件保存至:\n{save_path}\n\n{result_text}"
        except Exception as e:
            log(f"❌ Failed to write log file: {e}")
            return False, f"订单整合完成但日志保存失败:\n{str(e)}"

# ========== 增强版订单检查器 ==========
class EnhancedOrderChecker:
    """使用配置文件的订单检查器"""
    
    def __init__(self, config_manager=None):
        self.config_manager = config_manager
        # 新增：建立 full name/short name/email name normalize 映射到 short name
        self.fullname_to_short = {}
        if config_manager and hasattr(config_manager, 'config_path'):
            import pandas as pd
            import os
            config_path = config_manager.config_path if hasattr(config_manager, 'config_path') else None
            if config_path and os.path.exists(config_path):
                try:
                    df = pd.read_excel(config_path, sheet_name=None)
                    outlet_df = None
                    for key in df.keys():
                        if key.strip().lower() == "outlet":
                            outlet_df = df[key]
                            break
                    if outlet_df is not None:
                        for _, row in outlet_df.iterrows():
                            short = str(row.get("Short Name", "")).strip()
                            full = str(row.get("Outlet Full Name", "")).strip()
                            email_name = str(row.get("Name in Email", "")).strip()
                            for n in [short, full, email_name]:
                                n_norm = self._normalize(n)
                                if n and n_norm:
                                    self.fullname_to_short[n_norm] = short
                except Exception as e:
                    print("[EnhancedOrderChecker] Outlet mapping read error:", e)

    @staticmethod
    def _normalize(text):
        """标准化文本"""
        import re
        if not text:
            return ""
        # 只去除空格，保留括号内容以区分不同供应商
        return re.sub(r'[\s]', '', str(text).lower())

    def get_outlet_shortname(self, f5_value):
        """智能获取分店简称，优先用 config mapping"""
        import re
        
        if not f5_value or not isinstance(f5_value, str):
            return f"[EMPTY] {f5_value}"
        
        # 先嘗試完整匹配
        n = self._normalize(f5_value)
        if n in self.fullname_to_short:
            return self.fullname_to_short[n]
        
        # 處理包含配送日期資訊的 F5 值
        # 例如: "Sushi Express West Mall (MON,WED,FRI,SAT)" -> "Sushi Express West Mall"
        # 但保留包含門市代碼的括號，如 "Sushi Takeout CityVibe (GTM)"
        delivery_pattern = r'\s*\([A-Z]{2,4},[A-Z]{2,4},[A-Z]{2,4},[A-Z]{2,4}\)\s*$'
        cleaned_f5 = re.sub(delivery_pattern, '', f5_value.strip())
        
        if cleaned_f5 != f5_value:
            n = self._normalize(cleaned_f5)
            if n in self.fullname_to_short:
                return self.fullname_to_short[n]
        
        # fallback: 原有 hardcode/正则逻辑（可选）
        # ... existing code ...
        return f"[UNKNOWN] {f5_value}"

    def run_checklist(self, folder, log_callback=None, as_table=False):
        date_validator = None
        if self.config_manager:
            delivery_schedules = self.config_manager.delivery_schedule
            if delivery_schedules:
                date_validator = DeliveryDateValidator()
                date_validator.schedule = defaultdict(dict)
                for row in delivery_schedules:
                    supplier = row['supplier']
                    outlet = row['outlet_code']
                    days = DeliveryDateValidator().parse_delivery_days(row['delivery_days'])
                    if outlet == "ALL":
                        date_validator.schedule[supplier]['*'] = days
                    else:
                        date_validator.schedule[supplier][outlet] = days
        supplier_keywords = {}
        if self.config_manager:
            for supplier in self.config_manager.suppliers:
                supplier_keywords[supplier["name"]] = [supplier["name"].lower()]
        must_have_outlets = {}
        if self.config_manager:
            for supplier in self.config_manager.suppliers:
                must_have_outlets[supplier["name"]] = self.config_manager.get_required_outlets(supplier["name"])
        try:
            files = [f for f in os.listdir(folder) if f.endswith(".xlsx") and not f.startswith("~$")]
            def file_key(f):
                import re
                base = f.split("_Week_")[0]
                return self._normalize(base)
            normalized_files = {file_key(f): f for f in files}
            output = []
            table = []
            for supplier, keywords in supplier_keywords.items():
                matches = self._find_supplier_file(normalized_files, keywords)
                if not matches:
                    if as_table:
                        table.append({
                            "supplier": supplier,
                            "outlet": "-",
                            "cover_status": "❌",
                            "remark": "Supplier file not found"
                        })
                    else:
                        output.append(f"\n❌ {supplier} - Supplier file not found.")
                    continue
                for match in matches:
                    result, table_rows = self._process_supplier_file(
                        folder, match, must_have_outlets[supplier], 
                        date_validator, log_callback, as_table=True, supplier=supplier
                    )
                if as_table:
                    table.extend(table_rows)
                else:
                    output.extend(result)
            if as_table:
                return table
            return "\n".join(output)
        except Exception as e:
            if as_table:
                return [{"supplier": "-", "outlet": "-", "cover_status": "❌", "remark": f"Error: {str(e)}"}]
            return f"❌ Error running checklist: {str(e)}\n{traceback.format_exc()}"

    @classmethod
    def _find_supplier_file(cls, normalized_files, keywords):
        matches = []
        for k in keywords:
            nk = cls._normalize(k)
            for nf, of in normalized_files.items():
                if nf == nk:
                    matches.append(of)
        if matches:
            return matches
        # 不再做唯一模糊匹配，避免误配
        return None

    def _process_supplier_file(self, folder, filename, required_outlets, date_validator=None, log_callback=None, as_table=False, supplier=None):
        import os
        from openpyxl import load_workbook
        import tkinter.messagebox as messagebox
        output = []
        table = []
        found = set()
        unidentified = []
        date_errors = []
        unknown_f5 = set()
        try:
            wb = load_workbook(os.path.join(folder, filename), data_only=True)
            for s in wb.sheetnames:
                # 跳过 sheet 名为 'Sheet' 或空白的 sheet
                if not s.strip() or s.strip().lower() == 'sheet':
                    continue
                try:
                    f5 = wb[s]["F5"].value
                    code = self.get_outlet_shortname(f5)
                    sheet_name = s.strip().upper()
                    short_name_mismatch = False
                    
                    # 處理特殊情況：如果 F5 映射結果與工作表名稱不同，但都指向同一個門市
                    # 例如：F5→BUGIS, Sheet→BJ，但都指向 SushiPlus Bugis
                    if code != sheet_name and "[UNKNOWN]" not in code and "[EMPTY]" not in code:
                        # 檢查是否都映射到同一個門市
                        f5_normalized = self._normalize(f5) if f5 else ""
                        sheet_normalized = self._normalize(sheet_name)
                        
                        # 如果 F5 映射和工作表名稱都指向同一個門市，則不視為 mismatch
                        if f5_normalized in self.fullname_to_short and sheet_normalized in self.fullname_to_short:
                            if self.fullname_to_short[f5_normalized] == self.fullname_to_short[sheet_normalized]:
                                short_name_mismatch = False
                            else:
                                short_name_mismatch = True
                        else:
                            short_name_mismatch = True
                    if "[UNKNOWN]" in code or "[EMPTY]" in code:
                        unidentified.append((s, f5))
                        unknown_f5.add(str(f5).strip() if f5 else "[空白]")
                        if as_table:
                            table.append({
                                "supplier": supplier,
                                "outlet": s,
                                "cover_status": "⚠️",
                                "remark": f"F5 error: {f5}" + (f"; Short name mismatch: F5→{code}, Sheet→{sheet_name}" if short_name_mismatch else "")
                            })
                    else:
                        found.add(code)
                        date_status = "-"
                        remark = ""
                        if date_validator:
                            delivery_date = wb[s]['F8'].value if 'F8' in wb[s] else None
                            if delivery_date:
                                is_valid = date_validator.validate_order(
                                    supplier,
                                    code,
                                    delivery_date,
                                    log_callback
                                )
                                if not is_valid:
                                    date_status = "❌"
                                    remark = "Invalid delivery date"
                                    date_errors.append(code)
                                else:
                                    date_status = "✔️"
                            else:
                                date_status = "⚠️"
                                remark = ""
                        if short_name_mismatch:
                            if remark:
                                remark += "; "
                            remark += f"Short name mismatch: F5→{code}, Sheet→{sheet_name}"
                        if as_table:
                            table.append({
                                "supplier": supplier,
                                "outlet": code,
                                "cover_status": "✔️",
                                "remark": remark
                            })
                except Exception as e:
                    unidentified.append((s, f"[F5 error: {e}]") )
                    if as_table:
                        table.append({
                            "supplier": supplier,
                            "outlet": s,
                            "cover_status": "⚠️",
                            "remark": f"F5 error: {e} (sheet: {s}, file: {filename})"
                        })
            required = set([str(x).strip().upper() for x in required_outlets])
            found = set([str(x).strip().upper() for x in found])
            print(f"[DEBUG] supplier={supplier}")
            print(f"[DEBUG] required={required}")
            print(f"[DEBUG] found={found}")
            print(f"[DEBUG] missing={required - found}")
            print(f"[DEBUG] all sheet names in {filename}: {wb.sheetnames}")
            print(f"[DEBUG] unidentified sheets: {unidentified}")
            print(f"[DEBUG] unknown F5 values: {unknown_f5}")
            print(f"[DEBUG] F5 to code mappings:")
            for s in wb.sheetnames:
                if not s.strip() or s.strip().lower() == 'sheet':
                    continue
                try:
                    f5 = wb[s]["F5"].value
                    code = self.get_outlet_shortname(f5)
                    print(f"  Sheet: {s}, F5: {f5}, Code: {code}")
                except Exception as e:
                    print(f"  Sheet: {s}, F5 Error: {e}")
            
            # 檢查 missing outlet 是否有對應的 sheet 但 F5 和 sheet name 不一致
            missing_with_sheet_mismatch = set()
            for sheet_name in wb.sheetnames:
                if not sheet_name.strip() or sheet_name.strip().lower() == 'sheet':
                    continue
                sheet_name_upper = sheet_name.strip().upper()
                if sheet_name_upper in required and sheet_name_upper not in found:
                    try:
                        f5 = wb[sheet_name]["F5"].value
                        code = self.get_outlet_shortname(f5)
                        if code != sheet_name_upper and "[UNKNOWN]" not in code and "[EMPTY]" not in code:
                            missing_with_sheet_mismatch.add(sheet_name_upper)
                    except:
                        pass
            
            for o in sorted(required - found):
                if as_table:
                    if o in missing_with_sheet_mismatch:
                        # 有 sheet 但 F5 和 sheet name 不一致
                        table.append({
                            "supplier": supplier,
                            "outlet": o,
                            "cover_status": "⚠️",
                            "remark": f"Short name mismatch: Sheet→{o}"
                        })
                    else:
                        # 真正的 missing outlet
                        table.append({
                            "supplier": supplier,
                            "outlet": o,
                            "cover_status": "❌",
                            "remark": "Missing outlet"
                        })
            if as_table:
                # 检查结束后，弹窗提醒所有未能 mapping 的 F5 内容
                if unknown_f5:
                    msg = "以下 F5 内容未能自动 mapping 到 short name，请补充到 config 的 OUTLET sheet：\n" + "\n".join(f"- {f5val}" for f5val in sorted(unknown_f5))
                    try:
                        messagebox.showwarning("智能提示/Smart Reminder", msg)
                    except Exception:
                        print("[智能提示]", msg)
                return output, table
            # 原本文字报表
            output.append(f"\n=== {filename} ===")
            output.append(f"📊 Required: {len(required)}, Found: {len(found)}, Missing: {len(required - found)}")
            for o in sorted(required & found):
                output.append(f"✔️ {o}")
            for o in sorted(required - found):
                output.append(f"❌ {o}")
            for s, v in unidentified:
                output.append(f"⚠️ {s} => {v}")
            if unknown_f5:
                output.append("[智能提示] 以下 F5 内容未能自动 mapping 到 short name，请补充到 config 的 OUTLET sheet：")
                for f5val in sorted(unknown_f5):
                    output.append(f"  - {f5val}")
            return output, table
        except Exception as e:
            if as_table:
                return [{"supplier": "-", "outlet": "-", "cover_status": "❌", "remark": f"Error: {str(e)}"}]
            return f"❌ Error running checklist: {str(e)}\n{traceback.format_exc()}"

# ========== 增强版订单自动化 ==========
class EnhancedOrderAutomation(OrderAutomation):
    """支持邮件发送的订单自动化"""
    
    def __init__(self, config_manager=None):
        super().__init__()
        self.config_manager = config_manager
        self.email_sender = EmailSender(config_manager)
    
    def run_automation(self, source_folder, supplier_folder, 
                      log_callback=None, mapping_callback=None, 
                      email_callback=None):
        """运行自动化流程（仅 supplier 整合）"""
        success, result = super().run_automation(
            source_folder, supplier_folder, 
            log_callback=log_callback, mapping_callback=mapping_callback
        )
        if not success:
            return success, result
        supplier_files = []
        for file in os.listdir(supplier_folder):
            if file.endswith(".xlsx") and "Week" in file:
                supplier_name = file.split("_")[0]
                supplier_files.append({
                    "path": os.path.join(supplier_folder, file),
                    "supplier": supplier_name
                })
        if self.config_manager and email_callback:
            email_callback(supplier_files)
        return success, result

class YellowHighlightedOrderAutomation(OrderAutomation):
    """只整合有黃色標記的訂單"""
    
    def __init__(self, config_manager=None):
        super().__init__()
        self.config_manager = config_manager
    
    def has_yellow_highlight(self, ws, row, col):
        """檢查指定單元格是否有黃色標記"""
        try:
            cell = ws.cell(row=row, column=col)
            if cell.fill.start_color.rgb:
                fill_color = cell.fill.start_color.rgb
                print(f"[DEBUG] Cell ({row}, {col}) RGB: {fill_color}")
                
                # 檢查常見的黃色 RGB 值
                yellow_rgbs = ['FFFF00', 'FFFFFF00', 'FF00FF00', 'FFD966', 'FFF200', 'FFEB9C', 'FFE066', 'FFD700', 'FFE135', 'FFD800', 'FFE100']
                if fill_color in yellow_rgbs:
                    print(f"[DEBUG] ✅ 找到標準黃色: {fill_color}")
                    return True
                
                # 更寬鬆的黃色檢查：R 和 G 都很高，B 很低
                if fill_color.startswith('FF') and len(fill_color) == 8:
                    rgb = fill_color[2:]  # 去掉 alpha 通道
                    r = int(rgb[0:2], 16)
                    g = int(rgb[2:4], 16)
                    b = int(rgb[4:6], 16)
                    
                    # 黃色條件：R 和 G 都 > 240，B < 50
                    if r > 240 and g > 240 and b < 50:
                        print(f"[DEBUG] ✅ 找到寬鬆黃色: R={r}, G={g}, B={b}")
                        return True
                    
                    # 更寬鬆的條件：R 和 G 都 > 200，B < 100
                    if r > 200 and g > 200 and b < 100:
                        print(f"[DEBUG] ✅ 找到非常寬鬆黃色: R={r}, G={g}, B={b}")
                        return True
                
                print(f"[DEBUG] ❌ 不是黃色: {fill_color}")
            else:
                print(f"[DEBUG] Cell ({row}, {col}) 沒有填色")
            return False
        except Exception as e:
            print(f"[DEBUG] Error checking yellow highlight: {e}")
            return False
    
    def check_file_has_yellow_highlight(self, filepath, this_week_start, this_week_end):
        """檢查檔案是否有黃色標記的單元格"""
        try:
            print(f"[DEBUG] 開始檢查檔案: {filepath}")
            wb = openpyxl.load_workbook(filepath, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                print(f"[DEBUG] 檢查工作表: {sheet_name}")
                
                # 尋找日期行 - 使用本週日期範圍
                date_row_result = self.find_delivery_date_row(ws, this_week_start, this_week_end, file_path=filepath)
                if date_row_result is None or date_row_result[0] is None:
                    print(f"[DEBUG] 在工作表 {sheet_name} 中找不到日期行")
                    continue
                
                date_row = date_row_result[0]  # 取得行號
                print(f"[DEBUG] 找到日期行: {date_row}")
                
                # 檢查日期行下方的數量單元格是否有黃色標記
                for row in range(date_row + 1, min(date_row + 50, ws.max_row + 1)):
                    # 檢查是否為產品行（通常有產品名稱）- 使用 E 欄與 find_delivery_date_row 一致
                    product_name = ws.cell(row=row, column=5).value  # E 欄產品名稱
                    if not product_name or str(product_name).strip() == "":
                        continue
                    
                    print(f"[DEBUG] 檢查產品行: {row}, 產品: {product_name}")
                    
                    # 檢查日期欄位（通常是 F 到 L 欄）
                    for col in range(6, 13):  # F 到 L 欄
                        cell_value = ws.cell(row=row, column=col).value
                        if cell_value and (isinstance(cell_value, (int, float)) or 
                                         (isinstance(cell_value, str) and cell_value.replace('.', '').isdigit())):
                            print(f"[DEBUG] 檢查數量格: ({row}, {col}), 值: {cell_value}")
                            # 如果數量不為0且有黃色標記
                            if self.has_yellow_highlight(ws, row, col):
                                print(f"[DEBUG] ✅ 在檔案 {filepath} 中找到黃色標記!")
                                wb.close()
                                return True
            wb.close()
            print(f"[DEBUG] ❌ 在檔案 {filepath} 中沒有找到黃色標記")
            return False
        except Exception as e:
            print(f"[DEBUG] Error checking yellow highlight in {filepath}: {e}")
            return False
    
    def run_automation(self, source_folder, supplier_folder, log_callback=None, mapping_callback=None):
        from datetime import datetime, timedelta
        import os
        import openpyxl
        import xlwings as xw
        import traceback

        now = datetime.now()
        today = now.date()
        this_monday = today - timedelta(days=today.weekday())
        this_sunday = this_monday + timedelta(days=6)
        start_of_period = datetime.combine(this_monday, datetime.min.time())
        end_of_period = datetime.combine(this_sunday, datetime.max.time())
        save_path = os.path.join(source_folder, f"yellow_highlight_order_log_{now.strftime('%Y%m%d_%H%M%S')}.txt")
        log_lines = [
            f"🎯 Amendment Order Integration Log\n",
            f"Scan Start Time: {now}\n",
            f"Target Week: {start_of_period.date()} to {end_of_period.date()}\n"
        ]

        def log(message, include_timestamp=True):
            timestamp = datetime.now().strftime("%H:%M:%S") if include_timestamp else ""
            log_line = f"[{timestamp}] {message}" if include_timestamp else message
            log_lines.append(log_line + "\n")
            if log_callback:
                log_callback(log_line + "\n")

        files = [f for f in os.listdir(source_folder) if f.endswith((".xlsx", ".xls")) and not f.startswith("~$")]
        total_files = len(files)
        if total_files == 0:
            log("未找到Excel文件\nNo Excel files found in source folder. Exiting.")
            return False, "在源文件夹中未找到Excel文件"
        log(f"找到 {total_files} 个Excel文件\nFound {total_files} Excel files")
        log(f"目标周期: {start_of_period.strftime('%Y-%m-%d')} 到 {end_of_period.strftime('%Y-%m-%d')}")
        log(f"🔍 專注於本週訂單，忽略其他週期")

        # 1. 找出所有有黃色標記的 (門市, 廠商) sheet
        supplier_to_outlets = {}  # supplier: set(outlet)
        for filename in files:
            filepath = os.path.join(source_folder, filename)
            try:
                wb = openpyxl.load_workbook(filepath, data_only=True)
                for sheetname in wb.sheetnames:
                    ws = wb[sheetname]
                    # 跳過隱藏 sheet
                    if hasattr(ws, 'sheet_state') and ws.sheet_state != "visible":
                        continue
                    found_yellow = False
                    for row in ws.iter_rows():
                        for cell in row:
                            fill = cell.fill
                            if fill and fill.fgColor and hasattr(fill.fgColor, 'rgb'):
                                rgb = fill.fgColor.rgb
                                if isinstance(rgb, str) and (rgb.upper() == 'FFFFFF00' or rgb.upper() == 'FFFF00'):
                                    found_yellow = True
                                    break
                        if found_yellow:
                            break
                    if found_yellow:
                        supplier_to_outlets.setdefault(sheetname, set()).add(filepath)
                        log(f"✅ {filename} 發現黃色標記: {sheetname}")
                wb.close()
            except Exception as e:
                log(f"❌ 檢查 {filename} 發生錯誤: {e}")
        log("")
        log(f"📊 統計結果:")
        log(f"   - 總檔案數: {len(files)}")
        log(f"   - 有黃色標記的廠商: {len(supplier_to_outlets)}")
        log("")
        if not supplier_to_outlets:
            log("❌ 沒有找到任何有黃色標記的廠商")
            return False, "沒有找到任何有黃色標記的廠商"
        log(f"📋 準備整合以下廠商:")
        for supplier in supplier_to_outlets:
            log(f"   - {supplier}")
        log("")
        # 2. 以廠商為單位合併所有有黃色標記的門市的該廠商 sheet（用 xlwings 複製 sheet，保留格式/公式）
        supplier_files = []
        for supplier, filelist in supplier_to_outlets.items():
            try:
                supplier_path = os.path.join(supplier_folder, f"{supplier}_Amendment_{now.isocalendar()[1]}.xlsx")
                # 先建立空檔案
                if not os.path.exists(supplier_path):
                    from openpyxl import Workbook
                    wb = Workbook()
                    wb.save(supplier_path)
                app = xw.App(visible=False)
                try:
                    wb_dest = app.books.open(supplier_path)
                    for src_file in filelist:
                        try:
                            wb_src = app.books.open(src_file)
                            if supplier in [sht.name for sht in wb_src.sheets]:
                                sht = wb_src.sheets[supplier]
                                dest_sheet_name = os.path.splitext(os.path.basename(src_file))[0]
                                # 刪除同名分頁
                                for s in wb_dest.sheets:
                                    if s.name == dest_sheet_name:
                                        s.delete()
                                sht.api.Copy(Before=wb_dest.sheets[0].api)
                                wb_dest.sheets[0].name = dest_sheet_name
                            wb_src.close()
                        except Exception as e:
                            log(f"    ❌ Failed to copy {supplier} in {os.path.basename(src_file)}: {str(e)}\n{traceback.format_exc()}")
                    wb_dest.save()
                    wb_dest.close()
                    supplier_files.append(os.path.basename(supplier_path))
                    log(f"  ✅ 已整合: {os.path.basename(supplier_path)}")
                finally:
                    app.quit()
            except Exception as e:
                log(f"❌ 整合 {supplier} 發生錯誤: {e}\n{traceback.format_exc()}")
        result_text = f"\n🎯 本週訂單整合完成 / Amendment Order Integration Complete:\n"
        result_text += f"📦 已處理供應商文件: {len(supplier_files)}\n"
        result_text += f"📅 目標週期: {this_monday.strftime('%Y-%m-%d')} 到 {this_sunday.strftime('%Y-%m-%d')}\n\n"
        result_text += "📋 供應商文件列表:\n" + "\n".join([f"  ✅ {file}" for file in supplier_files])
        log(result_text)
        try:
            with open(save_path, "w", encoding="utf-8") as logfile:
                logfile.writelines(log_lines)
            log(f"\nLog saved at: {save_path}")
            return True, f"Amendment Order 整合完成！\n\n日志文件保存至:\n{save_path}\n\n{result_text}"
        except Exception as e:
            log(f"❌ Failed to write log file: {e}")
            return False, f"Amendment Order 整合完成但日志保存失败:\n{str(e)}"

# ========== 主应用程序 ==========
class SushiExpressApp(ctk.CTk):
    """主应用程序"""
    
    def __init__(self):
        super().__init__()
        self.title(f"Sushi Express Automation Tool v{VERSION}")
        self.geometry("1400x900")
        self.minsize(1200,800)
        self.configure(fg_color=DARK_BG)
        self.iconbitmap(resource_path("SELOGO22 - 01.ico"))
        self.protocol("WM_DELETE_WINDOW", self._on_close)
        # 初始化所有用到的 StringVar
        self.download_folder_var = ctk.StringVar()
        self.config_file_var = ctk.StringVar()
        self.checklist_folder_var = ctk.StringVar()
        self.master_config_var = ctk.StringVar()
        self.folder_vars = {}
        self.master_file_var = ctk.StringVar()
        self.output_folder_var = ctk.StringVar()
        self.email_supplier_folder_var = ctk.StringVar()
        self.email_master_config_var = ctk.StringVar()
        self.progress_popup = None
        self.mapping_popup = None
        self.outlet_config_var = ctk.StringVar()
        self.email_dialogs = []
        self.current_function = None
        self.nav_buttons = {}
        self.selected_outlook_account_idx = None
        self.checklist_search_var = ctk.StringVar()
        self._setup_ui()
        self.show_login()
    
    def _setup_ui(self):
        self.main_container = ctk.CTkFrame(self, fg_color="transparent")
        self.main_container.pack(fill="both", expand=True, padx=20, pady=20)

        # 左側導航欄
        self.nav_frame = ctk.CTkFrame(
            self.main_container,
            fg_color=DARK_PANEL,
            corner_radius=24,
            width=300
        )
        self.nav_frame.pack(side="left", fill="y", padx=(0, 10), pady=10)
        self.nav_frame.pack_propagate(False)

        # 加入 LOGO
        logo_img = load_image(LOGO_PATH, max_size=(220, 80))
        if logo_img:
            logo_label = ctk.CTkLabel(self.nav_frame, image=logo_img, text="")
            logo_label.image = logo_img  # 防止被垃圾回收
            logo_label.pack(pady=(18, 8))

        nav_title = ctk.CTkLabel(
            self.nav_frame,
            text="功能菜单\nFunction Menu",
            font=FONT_TITLE,
            text_color=ACCENT_BLUE,
            justify="center"
        )
        nav_title.pack(pady=20)
        ctk.CTkFrame(self.nav_frame, height=2, fg_color=ACCENT_BLUE).pack(fill="x", padx=20, pady=10)

        # 按鈕容器
        self.button_container = ctk.CTkFrame(self.nav_frame, fg_color="transparent")
        self.button_container.pack(fill="both", expand=True, padx=10, pady=10)

        # 右側內容區
        self.content_container = ctk.CTkFrame(
            self.main_container,
            fg_color=DARK_PANEL,
            corner_radius=24
        )
        self.content_container.pack(side="right", fill="both", expand=True, padx=10, pady=10)
        self.content_container.pack_propagate(False)

        # 右側內容區的標題區域
        self.content_header = ctk.CTkFrame(self.content_container, fg_color="transparent")
        self.content_header.pack(fill="x", padx=20, pady=20)

        self.function_title = ctk.CTkLabel(
            self.content_header, 
            text="", 
            font=FONT_TITLE,
            text_color=ACCENT_BLUE
        )
        self.function_title.pack(side="left")
        
        self.function_subtitle = ctk.CTkLabel(
            self.content_header, 
            text="", 
            font=FONT_SUB,
            text_color=TEXT_COLOR
        )
        self.function_subtitle.pack(side="left", padx=20)
        
        # 右侧内容区的主体
        self.content_body = ctk.CTkFrame(self.content_container, fg_color="transparent")
        self.content_body.pack(fill="both", expand=True, padx=20, pady=20)
        
        # 添加返回主菜单按钮
        back_frame = ctk.CTkFrame(self.content_container, fg_color="transparent")
        back_frame.pack(fill="x", padx=20, pady=10)
        ctk.CTkButton(
            back_frame, 
            text="返回主菜单\nBack to Main Menu", 
            command=self.show_main_menu,
            fg_color=ACCENT_PURPLE,
            hover_color=BTN_HOVER
        ).pack(side="right")
    
    def _on_close(self):
        print("on_close called")
        self.destroy()
    
    def show_login(self):
        """显示登录界面"""
        for w in self.content_container.winfo_children():
            if w != self.content_header and w != self.content_body:
                w.destroy()
        
        # 清空内容区
        for w in self.content_body.winfo_children():
            w.destroy()
        
        # 创建登录界面
        login_frame = ctk.CTkFrame(
            self.content_body, 
            fg_color=DARK_PANEL, 
            corner_radius=24,
            width=450, 
            height=400
        )
        login_frame.place(relx=0.5, rely=0.5, anchor="center")
        
        self.pwd_entry = ctk.CTkEntry(
            login_frame, 
            show="*", 
            font=FONT_MID,  # 改成較小字體
            width=300, 
            placeholder_text=t("password"), 
            fg_color=ENTRY_BG, 
            height=45
        )
        self.pwd_entry.pack(pady=(40,20))
        self.pwd_entry.bind("<Return>", lambda e: self._try_login())
        
        ctk.CTkButton(
            login_frame, 
            text=t("login_btn"), 
            command=self._try_login, 
            width=200,
            font=FONT_BIGBTN
        ).pack(pady=(0,20))
        
        ctk.CTkLabel(
            login_frame, 
            text=f"Version {VERSION} | {DEVELOPER}", 
            font=FONT_MID,
            text_color="#64748B"
        ).pack(side="bottom", pady=10)
        
        # 设置标题
        self.function_title.configure(text="系统登录\nSystem Login", font=FONT_TITLE)
        self.function_subtitle.configure(text="请输入密码進入系统\nPlease enter password to access the system", font=FONT_MID)
    
    def _try_login(self):
        """尝试登录"""
        if self.pwd_entry.get() == PASSWORD:
            self.show_main_menu()
        else:
            messagebox.showerror(t("error"), t("incorrect_pw"))
    
    def show_main_menu(self):
        # 清空内容区
        for w in self.content_body.winfo_children():
            w.destroy()

        # 设置标题
        self.function_title.configure(text=t("main_title"))
        self.function_subtitle.configure(text=t("select_function"))

        # 创建欢迎界面
        welcome_frame = ctk.CTkFrame(self.content_body, fg_color="transparent")
        welcome_frame.pack(fill="both", expand=True, padx=50, pady=50)

        welcome_text = (
            "欢迎使用Sushi Express 自动化工具\n"
            "Welcome to Sushi Express Automation Tool\n\n"
            "请从左侧菜单选择您需要的功能\n"
            "Please select a function from the left menu"
        )

        ctk.CTkLabel(
            welcome_frame,
            text=welcome_text,
            font=FONT_MID,
            text_color=TEXT_COLOR,
            justify="center"
        ).pack(expand=True)

        # **每次都重建按鈕**
        self._create_navigation_buttons()
    
    def _create_navigation_buttons(self):
        for widget in self.button_container.winfo_children():
            widget.destroy()
        inner_frame = ctk.CTkFrame(self.button_container, fg_color="transparent")
        inner_frame.pack(expand=True)
        functions = [
            ("download_title", self.show_download_ui, ACCENT_BLUE),
            ("automation_title", self.show_automation_ui, ACCENT_GREEN),
            ("checklist_title", self.show_checklist_ui, ACCENT_PURPLE),
            ("send_emails", self.show_email_sending_ui, ACCENT_RED),
            ("operation_supplies", self.show_operation_supplies_ui, "#F97316"),
            ("exit_system", self._on_close, "#64748B")
        ]
        for func_key, command, color in functions:
            btn = NavigationButton(
                inner_frame,
                text=t(func_key),
                command=command if func_key == "exit_system" else lambda c=command, k=func_key: self._select_function(c, k),
                fg_color=DARK_PANEL,
                anchor="center",
                border_width=2,
                border_color=ACCENT_BLUE,
                corner_radius=10,
                font=FONT_BIGBTN,
                text_color=get_contrast_color(DARK_PANEL),
                height=50
            )
            btn.pack(fill="x", pady=7, padx=10)
            self.nav_buttons[func_key] = btn
        # 新增：用戶指南按鈕
        help_btn = NavigationButton(
            inner_frame,
            text="用户指南\nUser Guide",
            command=self.show_user_guide,
            fg_color=ACCENT_PURPLE,
            anchor="center",
            border_width=2,
            border_color=ACCENT_PURPLE,
            corner_radius=10,
            font=FONT_BIGBTN,
            text_color=get_contrast_color(ACCENT_PURPLE),
            height=50
        )
        help_btn.pack(fill="x", pady=7, padx=10)
    
    def _select_function(self, command, func_key):
        """选择功能"""
        # 取消之前选中的按钮
        if self.current_function:
            self.nav_buttons[self.current_function].deselect()
        
        # 选中当前按钮
        self.nav_buttons[func_key].select()
        self.current_function = func_key
        
        # 执行功能
        command()
    
    def _show_function_ui(self, title_key, subtitle_key, content_callback):
        """显示功能界面"""
        # 清空内容区
        for w in self.content_body.winfo_children():
            w.destroy()
        
        # 设置标题
        self.function_title.configure(text=t(title_key))
        # 副標題統一放大
        if isinstance(subtitle_key, tuple):
            text = subtitle_key[0]
        else:
            text = t(subtitle_key)
        self.function_subtitle.configure(text=text, font=FONT_TITLE)
        
        # 创建内容框架
        content_frame = ctk.CTkFrame(self.content_body, fg_color="transparent")
        content_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # 构建功能UI
        content_callback(content_frame)
    
    # 下面是各个功能界面的修改，只需要将原来的show_xxx_ui方法改为使用_show_function_ui
    
    def show_download_ui(self):
        """显示下载界面"""
        def build(c):
            self.download_folder_var = ctk.StringVar()
            self.config_file_var = ctk.StringVar()
            
            # 创建表单框架
            form_frame = ctk.CTkFrame(c, fg_color="transparent")
            form_frame.pack(fill="both", expand=True, padx=50, pady=20)
            
            # 下载文件夹
            row1 = ctk.CTkFrame(form_frame)
            row1.pack(fill="x", pady=15)
            ctk.CTkLabel(row1, text="下载文件夹\nDownload Folder:", font=FONT_MID, anchor="w", justify="left").pack(side="left", padx=10)
            ctk.CTkEntry(row1, textvariable=self.download_folder_var, font=FONT_MID, state="readonly", width=400).pack(side="left", expand=True, fill="x", padx=5)
            # 美化的瀏覽按鈕
            browse_btn = ctk.CTkButton(
                row1, 
                text="浏览...\nBrowse...", 
                font=("Microsoft YaHei", 11, "bold"),
                command=self._select_download_folder,
                corner_radius=8,
                hover_color="#1976d2",
                height=30
            )
            browse_btn.pack(side="left", padx=5)
            
            # 分店配置文件
            row2 = ctk.CTkFrame(form_frame)
            row2.pack(fill="x", pady=15)
            ctk.CTkLabel(row2, text="分店配置文件（可选）\nOutlet Config (Optional):", font=FONT_MID, anchor="w", justify="left").pack(side="left", padx=10)
            ctk.CTkEntry(row2, textvariable=self.config_file_var, font=FONT_MID, state="readonly", width=400).pack(side="left", expand=True, fill="x", padx=5)
            # 美化的瀏覽按鈕
            browse_btn2 = ctk.CTkButton(
                row2, 
                text="浏览...\nBrowse...", 
                font=("Microsoft YaHei", 11, "bold"),
                command=self._select_config_file,
                corner_radius=8,
                hover_color="#1976d2",
                height=30
            )
            browse_btn2.pack(side="left", padx=5)
            # 开始下载按钮
            btn_frame = ctk.CTkFrame(c, fg_color="transparent")
            btn_frame.pack(pady=30)
            
            # 主要下載按鈕（已整合避免重複功能）
            GlowButton(
                btn_frame, 
                text="開始下載\nStart Download",
                command=self._run_download,
                glow_color=ACCENT_BLUE
            ).pack(fill="x", expand=True, padx=10, pady=5)
            
            # Amendment 下載按鈕
            GlowButton(
                btn_frame, 
                text="下載 Amendment\nDownload Amendments",
                command=self._run_download_amendments,
                glow_color="#ef4444"
            ).pack(fill="x", expand=True, padx=10, pady=5)
            # 确保"查看邮件内容"按钮始终显示，且更大
            def show_extracted_bodies():
                from tkinter import messagebox
                import os
                import tkinter as tk
                folder = self.download_folder_var.get()
                week_no = datetime.now().isocalendar()[1]
                save_path = os.path.join(folder, f"Week_{week_no}")
                log_file = os.path.join(save_path, "email_bodies_log.txt")
                bodies = getattr(OutlookDownloader, 'extracted_bodies', [])
                if not bodies and os.path.exists(log_file):
                    with open(log_file, "r", encoding="utf-8") as f:
                        content = f.read()
                    bodies = content.split("\n\n——— 邮件 ") if content else []
                    if bodies and not bodies[0].startswith("——— 邮件 "):
                        bodies[0] = "——— 邮件 1 ———\n" + bodies[0]
                    bodies = [b if b.startswith("——— 邮件 ") else "——— 邮件 " + b for b in bodies]
                if not bodies:
                    messagebox.showinfo("无内容", "请先下载邮件，再查看邮件内容！")
                    return
                # 搜索功能
                win = tk.Toplevel(self)
                win.title("邮件内容提取结果/Extracted Email Bodies")
                win.geometry("900x700")
                search_var = tk.StringVar()
                def filter_bodies(keyword):
                    keyword = keyword.lower().strip()
                    if not keyword:
                        return bodies
                    result = []
                    for b in bodies:
                        if keyword in b.lower():
                            result.append(b)
                    return result
                # 美化顯示
                from tkinter import scrolledtext
                text_widget = scrolledtext.ScrolledText(win, wrap="word", font=("Microsoft JhengHei", 15), bg="#f1f5f9", fg="#22292f")
                text_widget.pack(fill="both", expand=True, padx=10, pady=(50,10))
                def update_display():
                    filtered = filter_bodies(search_var.get())
                    pretty = []
                    for idx, b in enumerate(filtered, 1):
                        lines = b.split("\n")
                        # 主題/發件人加粗，正文分隔線
                        pretty.append(f"\n{'='*40}\n")
                        for l in lines:
                            if l.startswith("[发件人]") or l.startswith("[主题]"):
                                pretty.append(f"{l}\n")
                            elif l.startswith("[内容]"):
                                pretty.append(f"\n{l}\n{'-'*30}\n")
                            else:
                                pretty.append(f"{l}\n")
                    pretty_text = "".join(pretty)
                    text_widget.delete("1.0", "end")
                    text_widget.insert("1.0", pretty_text)
                def on_search(*args):
                    update_display()
                search_var.trace_add("write", on_search)
                search_entry = tk.Entry(win, textvariable=search_var, font=("Microsoft JhengHei", 14))
                search_entry.place(x=10, y=10, width=400, height=32)
                search_entry.insert(0, "输入关键字搜索/Type to search...")
                def on_focus_in(event):
                    if search_entry.get() == "输入关键字搜索/Type to search...":
                        search_entry.delete(0, "end")
                search_entry.bind("<FocusIn>", on_focus_in)
                # 支持复制
                def copy_all():
                    import pyperclip
                    pretty = text_widget.get("1.0", "end-1c")
                    pyperclip.copy(pretty)
                    messagebox.showinfo("复制成功", "已复制全部邮件内容！")
                # 美化 Copy All 按鈕（小巧、圓角、亮綠、hover 深綠、字體適中）
                def on_enter(e):
                    btn.config(bg="#16a34a")
                def on_leave(e):
                    btn.config(bg="#22c55e")
                btn = tk.Button(
                    win,
                    text="复制全部内容\nCopy All",
                    command=copy_all,
                    bg="#22c55e",
                    fg="#fff",
                    font=("Microsoft JhengHei", 16, "bold"),
                    relief="ridge",
                    bd=2,
                    activebackground="#16a34a",
                    activeforeground="#fff",
                    cursor="hand2",
                    highlightthickness=1,
                    highlightbackground="#22d3ee"
                )
                btn.place(x=420, y=10, width=140, height=32)
                btn.bind("<Enter>", on_enter)
                btn.bind("<Leave>", on_leave)
                # 調整正文字體為14
                text_widget.config(font=("Microsoft JhengHei", 14))
                update_display()
            GlowButton(
                btn_frame,
                text="查看邮件内容\nCheck Email Bodies",
                command=show_extracted_bodies,
                width=300,
                height=48,
                glow_color=ACCENT_GREEN
            ).pack(pady=10)
        
        self._show_function_ui(
            "download_title", 
            "download_desc", 
            build
        )
    
    def show_checklist_ui(self):
        """显示检查表界面"""
        def build(c):
            # 创建表单框架
            self.checklist_folder_var = ctk.StringVar()
            self.master_config_var = ctk.StringVar()
            # 创建表单框架
            form_frame = ctk.CTkFrame(c, fg_color="transparent")
            form_frame.pack(fill="both", expand=True, padx=50, pady=5)
            # 订单文件夹
            row1 = ctk.CTkFrame(form_frame)
            row1.pack(fill="x", pady=5)
            ctk.CTkLabel(row1, text="訂單文件夾\nOrder Folder:", font=FONT_BIGBTN).pack(side="left", padx=10)
            ctk.CTkEntry(row1, textvariable=self.checklist_folder_var, font=FONT_MID, state="readonly", width=250).pack(side="left", expand=True, fill="x", padx=5)
            checklist_browse_btn = ctk.CTkButton(
                row1, 
                text=str(t("browse")), 
                font=("Microsoft YaHei", 11, "bold"),
                command=self._select_checklist_folder,
                corner_radius=8,
                hover_color="#1976d2",
                height=30
            )
            checklist_browse_btn.pack(side="left", padx=5)
            # 统一配置文件
            row3 = ctk.CTkFrame(form_frame)
            row3.pack(fill="x", pady=5)
            ctk.CTkLabel(row3, text="統一配置文件\nMaster Config (Excel):", font=FONT_BIGBTN).pack(side="left", padx=10)
            ctk.CTkEntry(row3, textvariable=self.master_config_var, font=FONT_MID, state="readonly", width=250).pack(side="left", expand=True, fill="x", padx=5)
            checklist_config_browse_btn = ctk.CTkButton(
                row3, 
                text="瀏覽...\nBrowse...", 
                font=("Microsoft YaHei", 11, "bold"),
                command=lambda: self._select_config_file(self.master_config_var, [("Excel files", "*.xlsx")]),
                corner_radius=8,
                hover_color="#1976d2",
                height=30
            )
            checklist_config_browse_btn.pack(side="left", padx=5)
            # 搜索區（永遠顯示）
            search_frame = ctk.CTkFrame(c, fg_color="transparent")
            search_frame.pack(fill="x", pady=5)
            # 移除搜尋範圍下拉選單
            # if not hasattr(self, 'checklist_search_scope_var'):
            #     self.checklist_search_scope_var = ctk.StringVar(value="全部")
            # scope_options = ["全部", "供應商", "分店"]
            # scope_menu = ctk.CTkOptionMenu(search_frame, variable=self.checklist_search_scope_var, values=scope_options, width=90, font=FONT_MID)
            # scope_menu.pack(side="left", padx=4)
            # Supplier 下拉選單
            if not hasattr(self, 'checklist_supplier_filter_var'):
                self.checklist_supplier_filter_var = ctk.StringVar(value="全部")
            if not hasattr(self, 'checklist_outlet_filter_var'):
                self.checklist_outlet_filter_var = ctk.StringVar(value="全部")
            def get_unique_suppliers():
                return ["全部"] + sorted(list({row["supplier"] for row in getattr(self, '_checklist_table_data', []) if row["supplier"]}))
            def get_unique_outlets():
                return ["全部"] + sorted(list({row["outlet"] for row in getattr(self, '_checklist_table_data', []) if row["outlet"]}))
            ctk.CTkLabel(search_frame, text="Supplier", font=("Microsoft JhengHei", 16, "bold"), text_color="#3b82f6").pack(side="left", padx=(0,2))
            self.supplier_menu = ctk.CTkOptionMenu(
                search_frame, variable=self.checklist_supplier_filter_var, values=get_unique_suppliers(), width=140, font=("Microsoft JhengHei", 15, "bold"),
                fg_color="#3b82f6", button_color="#3b82f6", button_hover_color="#2563eb", text_color="#fff", corner_radius=16,
                command=lambda _: self._filter_checklist_table())
            self.supplier_menu.pack(side="left", padx=4)
            ctk.CTkLabel(search_frame, text="Outlet", font=("Microsoft JhengHei", 16, "bold"), text_color="#3b82f6").pack(side="left", padx=(10,2))
            self.outlet_menu = ctk.CTkOptionMenu(
                search_frame, variable=self.checklist_outlet_filter_var, values=get_unique_outlets(), width=140, font=("Microsoft JhengHei", 15, "bold"),
                fg_color="#3b82f6", button_color="#3b82f6", button_hover_color="#2563eb", text_color="#fff", corner_radius=16,
                command=lambda _: self._filter_checklist_table())
            self.outlet_menu.pack(side="left", padx=4)
            ctk.CTkLabel(search_frame, text="搜索/Filter:", font=("Microsoft JhengHei", 15, "bold"), text_color="#3b82f6").pack(side="left", padx=6)
            if not hasattr(self, 'checklist_search_var'):
                self.checklist_search_var = ctk.StringVar()
            search_entry = ctk.CTkEntry(search_frame, textvariable=self.checklist_search_var, font=("Microsoft JhengHei", 14), width=180, corner_radius=10, border_width=2, border_color="#3b82f6")
            search_entry.pack(side="left", padx=4)
            search_entry.bind("<KeyRelease>", lambda e: self._filter_checklist_table())
            def set_search_keyword(keyword):
                self.checklist_search_var.set(keyword)
                self._filter_checklist_table()
            ctk.CTkButton(
                search_frame, text="Missing", width=90, height=34, font=("Microsoft JhengHei", 15, "bold"),
                fg_color="#2563eb", hover_color="#60a5fa", text_color="#fff", corner_radius=16, border_width=2, border_color="#fff",
                command=lambda: set_search_keyword("missing")
            ).pack(side="left", padx=6)
            ctk.CTkButton(
                search_frame, text="Mismatch", width=100, height=34, font=("Microsoft JhengHei", 15, "bold"),
                fg_color="#f59e42", hover_color="#fbbf24", text_color="#fff", corner_radius=16, border_width=2, border_color="#fff",
                command=lambda: set_search_keyword("mismatch")
            ).pack(side="left", padx=6)
            # 表格區（永遠顯示，沒資料時顯示空表格）
            table_frame = ctk.CTkFrame(c, fg_color="transparent")
            table_frame.pack(fill="both", expand=True, padx=10, pady=10)
            import tkinter.ttk as ttk
            style = ttk.Style()
            style.theme_use('default')
            style.configure("Custom.Treeview", background="#1e293b", fieldbackground="#1e293b", foreground="#e2e8f0", rowheight=28, font=("Microsoft JhengHei", 12))
            style.configure("Custom.Treeview.Heading", background="#334155", foreground="#60a5fa", font=("Microsoft JhengHei", 13, "bold"))
            style.map("Custom.Treeview", background=[('selected', '#334155')])
            if not hasattr(self, 'checklist_table'):
                self.checklist_table = None
            self.checklist_table = ttk.Treeview(
                table_frame, 
                columns=("supplier", "outlet", "cover_status", "remark"), 
                show="headings", 
                height=8,
                style="Custom.Treeview"
            )
            col_labels = [
                ("supplier", "供應商/Supplier"),
                ("outlet", "分店/Outlet"),
                ("cover_status", "覆蓋狀態/Cover"),
                ("remark", "備註/Remark")
            ]
            for col, label in col_labels:
                self.checklist_table.heading(col, text=label)
                self.checklist_table.column(col, width=140 if col!="remark" else 300, anchor="center")
            self.checklist_table.pack(fill="both", expand=True)
            self.checklist_table.bind("<Double-1>", self._on_checklist_row_double_click)
            # 複製/匯出按鈕（永遠顯示）
            btn_frame = ctk.CTkFrame(c, fg_color="transparent")
            btn_frame.pack(pady=10)
            btns = [
                ("匯出Excel\nExport Excel", self._export_checklist_table, "#10b981"),
                ("查看必要門市\nView Required Outlets", self._show_required_outlets_window, "#8b5cf6"),
                ("交叉檢查\nCross Check", self._run_cross_check_email_log, "#f59e42"),
            ]
            for txt, cmd, color in btns:
                GlowButton(
                    btn_frame,
                    text=txt,
                    command=cmd,
                    width=90,
                    height=28,
                    glow_color=color
                ).pack(side="left", padx=4, pady=2)
            for col in range(2):
                btn_frame.grid_columnconfigure(col, weight=1)
            # Run Check 按鈕（大、中文在上英文在下、置中、永遠顯示）
            run_btn_frame = ctk.CTkFrame(c, fg_color="transparent")
            run_btn_frame.pack(fill="x", pady=15)
            GlowButton(
                run_btn_frame,
                text="執行檢查\nRun Check",
                command=self._run_enhanced_checklist,
                width=200,
                height=56,
                glow_color="#a78bfa"
            ).pack(anchor="center")
            # 初始化表格數據
            if not hasattr(self, '_checklist_table_data'):
                self._checklist_table_data = []
            self._refresh_checklist_table()
        
        self._show_function_ui(
            "checklist_title", 
            "checklist_desc", 
            build
        )

# ========== 入口点 ==========
if __name__ == '__main__':
    try:
        app = SushiExpressApp()
        app.mainloop()
    except Exception as e:
        messagebox.showerror("Error", f"Startup failed: {e}")
        sys.exit(1)
